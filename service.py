import time
import os
import openpyxl
import requests
import re
import telebot
import sys
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pathlib import Path
from datetime import datetime
from requests_toolbelt.multipart.encoder import MultipartEncoder
from urllib.parse import urlencode
from bs4 import BeautifulSoup
from requests.exceptions import ConnectionError, ReadTimeout
# Load các biến môi trường từ file .env
dotenv_path = Path('.env')
load_dotenv(dotenv_path=dotenv_path)


# Lấy thông tin đăng nhập từ biến môi trường
BOT_TOKEN = os.getenv('BOT_TOKEN')
USERNAME = os.getenv('VISA_USERNAME')
PASSWORD = os.getenv('VISA_PASSWORD')
bot = telebot.TeleBot(BOT_TOKEN)

# Định nghĩa các URL cần sử dụng
search_url = os.getenv('SEARCH_URL')
login_url = os.getenv('LOGIN_URL')
doOpenSuspend_url = os.getenv('DO_OPEN_SUSPEND_URL')
detail_url = os.getenv('DETAIL_URL')

# Khởi tạo một session HTTP sử dụng module requests
session = requests.Session()

# Gửi yêu cầu đăng nhập và lưu lại cookies
login_data = {
    'j_username': USERNAME,
    'j_password': PASSWORD
}

def login(username, user):
    try:
        encoded_login_data = urlencode(login_data)

        response = session.post(login_url, data=encoded_login_data, headers={
                                'Content-Type': 'application/x-www-form-urlencoded'})
        cookies = session.cookies.get_dict()

        cookies_str = "; ".join([f"{k}={v}" for k, v in cookies.items()])

        # Chuẩn bị dữ liệu và gửi yêu cầu tìm kiếm
        account_name = username
        m = 'getList'
        typeAction = "LIST"
        organizationId = "64"
        hierarchy = "on"
        search_data = {
            'bean.serviceInfo.account': account_name,
            'm': m,
            "typeAction": typeAction,
            "bean.organizationId": organizationId,
            "bean.hierarchy": hierarchy
        }

        multipart_data = MultipartEncoder(fields=search_data)

        response_search = session.post(search_url, data=multipart_data, headers={
            'Content-Type': multipart_data.content_type, 'Cookie': cookies_str})

        # Phân tích kết quả tìm kiếm để lấy thông tin chi tiết
        soup = BeautifulSoup(response_search.text, 'html.parser')

        first_a_tag = soup.find('a', href=re.compile(r'bean\.id=\d+'))
        if first_a_tag:
            checkAccount_ACTIVE = check_account_status(username)
            if checkAccount_ACTIVE == 'ACTIVE':
                return ("Tài khoản đã mở cước - Không thể thực hiện mở cước")

            href = first_a_tag['href']
            bean_id = re.search(r'bean\.id=(\d+)', href).group(1)

            url_detail = detail_url+bean_id

            response_search_detail = session.get(
                url_detail, headers={'Cookie': cookies_str})

            soup2 = BeautifulSoup(response_search_detail.text, 'html.parser')

            input_tag = soup2.find(
                'input', attrs={'name': 'bean.customer.code'})

            if input_tag:
                id_user = input_tag['value']

                description = "Bot telegram mở cước, User yêu cầu : {}".format(
                    user)

                openSuspend_data = {
                    "bean.id": bean_id,
                    "bean.customer.code": id_user,
                    "bean.description": description
                }

                multipart_openSuspend_data = MultipartEncoder(
                    fields=openSuspend_data)

                response_openSuspend = session.post(doOpenSuspend_url, data=multipart_openSuspend_data, headers={
                    'Content-Type': multipart_openSuspend_data.content_type, 'Cookie': cookies_str})

                checkAccount = check_account_status(username)
                if checkAccount == 'ACTIVE':
                    return ("Mở cước thành công")
                else:
                    return ("Mở cước không thành công")

            else:
                return ('Không tìm thấy tài khoản này')

        else:
            return ('Không tìm thấy tài khoản này')
    except:
        return ('Lỗi kết nối hãy thử lại')


def check_account_status(username):
    try:
        encoded_login_data = urlencode(login_data)

        response = session.post(login_url, data=encoded_login_data, headers={
                                'Content-Type': 'application/x-www-form-urlencoded'})
        cookies = session.cookies.get_dict()

        cookies_str = "; ".join([f"{k}={v}" for k, v in cookies.items()])

        # Chuẩn bị dữ liệu và gửi yêu cầu tìm kiếm
        account_name = username
        m = 'getList'
        typeAction = "LIST"
        organizationId = "64"
        hierarchy = "on"
        search_data = {
            'bean.serviceInfo.account': account_name,
            'm': m,
            "typeAction": typeAction,
            "bean.organizationId": organizationId,
            "bean.hierarchy": hierarchy
        }

        multipart_data = MultipartEncoder(fields=search_data)

        response_search = session.post(search_url, data=multipart_data, headers={
            'Content-Type': multipart_data.content_type, 'Cookie': cookies_str})

        soup = BeautifulSoup(response_search.text, 'html.parser')
        first_a_tag = soup.find('a', href=re.compile(r'bean\.id=\d+'))

        if first_a_tag:
            # Tìm các phần tử có id là "service_status_4"
            elements_ACTIVE = soup.find_all(id="service_status_1")
            elements_SUSPEND = soup.find_all(id="service_status_4")
            if elements_ACTIVE:
                return "ACTIVE"

            if elements_SUSPEND:
                return "SUSPEND"
        else:
            return ('Không tìm thấy tài khoản này')
    except:
        return ('Lỗi kết nối hãy thử lại')


@bot.message_handler(commands=['kt'])
def check_account_command(message):
    try:
        username = message.text.replace('/kt', '').strip()
        bot.reply_to(message, f"Đang thực hiện kiểm tra tài khoản : {username}")
        user = message.from_user.full_name
        print("bot")
        if USERNAME and PASSWORD:
            status = check_account_status(username)
            bot.reply_to(message, status)
    except:
        bot.reply_to(message, f"Đã xảy ra lỗi khi kiểm tra tài khoản")


@bot.message_handler(commands=['help'])
def help_command(message):
    try:
        key_string = """/mc Tên-tài-khoản : Mở cước
/kt Tên-tài-khoản : Kiểm tra cước
/dulieu : Xem nhật ký"""
        bot.reply_to(message, key_string)
    except Exception as e:
        bot.reply_to(message, "Mất kết nối")

@bot.message_handler(commands=['mc'])
def login_command(message):
    try:
        username = message.text.replace('/mc', '').strip()
        bot.reply_to(message, f"Đang thực hiện mở khóa tài khoản: {username}")
        user = message.from_user.full_name
        thoigian = message.date
        # print(username)

        if USERNAME and PASSWORD:
            status = login(username, user)

            bot.reply_to(message, status)

            formatted_time = datetime.fromtimestamp(
                thoigian).strftime('%Y-%m-%d %H:%M:%S')
            formatted_time = formatted_time.encode(
                'ascii', 'replace').decode('ascii')
            if message.chat.type == 'group' or message.chat.type == 'supergroup':
                group_name = message.chat.title
                formatted_data = f"{user}, {username}, {formatted_time}, {status},Group: {group_name}"
            else:
                formatted_data = f"{user}, {username} , {formatted_time}, {status},Chat trực tiếp với Bot"
            with open('du_lieu.txt', 'a', encoding='utf-8') as file:
                file.write(formatted_data + "\n")

            path_to_txt = 'du_lieu.txt'

            workbook = openpyxl.Workbook()

            sheet = workbook.active

            with open(path_to_txt, 'r', encoding='utf-8') as file:
                lines = file.readlines()
            for row, line in enumerate(lines, start=3):
                values = line.strip().split(',')
                for col, value in enumerate(values, start=1):
                    sheet.cell(row=row, column=col, value=value)
                    column_letter = get_column_letter(col)
                    sheet.column_dimensions[column_letter].auto_size = True

            title_cell = sheet['A1']
            title_NTT = sheet['A2']
            title_TK = sheet['B2']
            title_TG = sheet['C2']
            title_PH = sheet['D2']
            title_GP = sheet['E2']

            title_NTT.value = 'Người Thực Hiện'
            title_TK.value = 'Tài Khoản Mở Cước'
            title_TG.value = 'Thời Gian'
            title_PH.value = 'Phản Hồi'
            title_GP.value = 'Nơi gửi'

            font = Font(bold=True)
            alignment = Alignment(horizontal='center')

            title_NTT.font = font
            title_TK.font = font
            title_TG.font = font
            title_PH.font = font
            title_GP.font = font

            title_NTT.alignment = alignment
            title_TK.alignment = alignment
            title_TG.alignment = alignment
            title_PH.alignment = alignment
            title_GP.alignment = alignment

            title_cell.value = 'Bảng Thống Kê'
            title_cell.font = Font(size=13, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            sheet.merge_cells('A1:E1')

            # Lưu workbook thành file Excel
            desktop_path = os.path.expanduser("./")
            file_name = "du_lieu.xlsx"
            path_to_excel = os.path.join(desktop_path, file_name)
            workbook.save(path_to_excel)
    except:
        bot.reply_to(message, f"Đã xảy ra lỗi khi ghi dữ liệu vào log")


@bot.message_handler(commands=['dulieu'])
def send_data_command(message):
    try:
        path_to_excel = 'du_lieu.xlsx'

        with open(path_to_excel, 'rb') as file:
            bot.send_document(message.chat.id, file)

    except Exception as e:
        bot.reply_to(message, f"Có lỗi xảy ra: {str(e)}")



try:
    bot.infinity_polling(timeout=10, long_polling_timeout=5)
except (ConnectionError, ReadTimeout) as e:
    sys.stdout.flush()
    os.execv(sys.argv[0], sys.argv)
else:
    bot.infinity_polling(timeout=10, long_polling_timeout=5)
