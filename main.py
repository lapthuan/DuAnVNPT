from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from pathlib import Path
import telebot
from datetime import datetime
dotenv_path = Path('.env')
load_dotenv(dotenv_path=dotenv_path)


BOT_TOKEN = os.getenv('BOT_TOKEN')
USERNAME = os.getenv('VISA_USERNAME')
PASSWORD = os.getenv('VISA_PASSWORD')
bot = telebot.TeleBot(BOT_TOKEN)
options = webdriver.ChromeOptions()

options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
# options.add_experimental_option("detach", True)
# Biến kiểm tra đã đăng nhập hay chưa


def login(username,user):
    try:
        print('try unlock')
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        print('driver ok')
        driver.get('http://10.156.7.25/visa/login.vnpt')
        print('get visa fe ok')

        taikhoan = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td/div/table/tbody/tr/td[2]/div/div/form/table/tbody/tr[3]/td/input')
        taikhoan.send_keys(USERNAME)

        matkhau = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td/div/table/tbody/tr/td[2]/div/div/form/table/tbody/tr[5]/td/input')
        matkhau.send_keys(PASSWORD)

        dangnhap = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td/div/table/tbody/tr/td[2]/div/div/form/table/tbody/tr[6]/td/input')
        dangnhap.click()

   
        select = driver.find_element(By.CSS_SELECTOR, "span.title")
        hover = ActionChains(driver).move_to_element(select)
        hover.perform()
        print('hover ok')

        time.sleep(.5)

        tracuu = driver.find_element("xpath", '/html/body/div[1]/div[2]/header/div[1]/nav/ul/li[1]/ul/li[1]/a')
        tracuu.click()
        
        uservisa = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td[2]/div/fieldset/form/table/tbody/tr[10]/td[2]/input')
        uservisa.send_keys(username)
        
        timkiem = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td[2]/div/fieldset[1]/form/table/tbody/tr[18]/td[2]/input[1]')
        timkiem.click()


        time.sleep(.5)

        div_elementTimKiem = driver.find_element(By.CSS_SELECTOR, 'div.line.left.color1')

        textTimKiem = div_elementTimKiem.text

        if textTimKiem == "Tổng số kết quả trả về:  0":
            driver.close()
            return "Không tìm thấy tài khoản"

        time.sleep(.5)

        radio = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td[2]/div/fieldset[2]/table/tbody/tr/td[2]/input')
        radio.click()

        time.sleep(1)

        motamngung = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td[2]/div/div[2]/input[11]')
        if motamngung.is_displayed():
        
            motamngung.click()
            
            ghichu = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td[2]/div/form/fieldset[2]/table/tbody/tr/td[2]/textarea')
            ghichustring = "Bot telegram mở cước, User yêu cầu : {}".format(user)
            ghichu.send_keys(ghichustring)  
                
            thuchien = driver.find_element("xpath", '/html/body/div[1]/table/tbody/tr/td[2]/div/form/div/input[1]')
            thuchien.click()
            
            time.sleep(1)
            wait = WebDriverWait(driver, 10)
            wait.until(EC.alert_is_present())
           
            al = driver.switch_to.alert
            al.accept()

            time.sleep(1)

            driver.close()
            print('unlock ok ok')
            return "Mở cước thành công"
        else:
            driver.close()
            return "Tài khoản đang được kích hoạt"
      
    except:
        driver.close()
        return "Lỗi"
      

@bot.message_handler(commands=['mc'])
def login_command(message):
    print('login_command')
    try:
        username = message.text.replace('/mc', '').strip()
        bot.reply_to(message, f"Đang thực hiện mở khóa tài khoản: {username}")
        user = message.from_user.full_name
        thoigian = message.date
        # print(username)
       
        if USERNAME and PASSWORD:
            status = login(username,user)

            bot.reply_to(message, status)

            
            formatted_time = datetime.fromtimestamp(thoigian).strftime('%Y-%m-%d %H:%M:%S')
            formatted_time = formatted_time.encode('ascii', 'replace').decode('ascii')
            if message.chat.type == 'group' or message.chat.type == 'supergroup':
                group_name = message.chat.title
                formatted_data = f"Người thực hiện: {user}, Tài khoản mở cước: {username} , Thời gian: {formatted_time}, Phản hồi: {status}, Group: {group_name}"
            else:
                formatted_data = f"Người thực hiện: {user}, Tài khoản mở cước: {username} , Thời gian: {formatted_time}, Phản hồi: {status}, Bot" 
            with open('du_lieu.txt', 'a', encoding='utf-8') as file:
                file.write(formatted_data + "\n")

        
            path_to_txt = 'du_lieu.txt'

        
            workbook = openpyxl.Workbook()

        
            sheet = workbook.active

            with open(path_to_txt, 'r', encoding='utf-8') as file:
                lines = file.readlines()
            for row, line in enumerate(lines, start=3):             
                values = line.strip().split(',')
                for col, value in enumerate(values, start=1):
                    sheet.cell(row=row, column=col, value=value)
                    column_letter = get_column_letter(col)
                    sheet.column_dimensions[column_letter].auto_size = True

        
            title_cell = sheet['A1']
            title_NTT = sheet['A2']
            title_TK = sheet['B2']
            title_TG = sheet['C2']
            title_PH = sheet['D2']
            title_GP = sheet['E2']
            
            title_NTT.value = 'Người Thực Hiện'
            title_TK.value = 'Tài Khoản Mở Cước'
            title_TG.value = 'Thời Gian'
            title_PH.value = 'Phản Hồi'
            title_GP.value = 'Nơi Gửi'

            font = Font(bold=True)
            alignment = Alignment(horizontal='center')

            title_NTT.font = font
            title_TK.font = font
            title_TG.font = font
            title_PH.font = font
            title_GP.font = font

            title_NTT.alignment = alignment
            title_TK.alignment = alignment
            title_TG.alignment = alignment
            title_PH.alignment = alignment
            title_GP.alignment = alignment

            title_cell.value = 'Bảng Thống Kê'
            title_cell.font = Font(size=13, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            sheet.merge_cells('A1:E1')

            # Lưu workbook thành file Excel
            desktop_path = os.path.expanduser("~/Desktop")
            file_name = "du_lieu.xlsx"
            path_to_excel = os.path.join(desktop_path, file_name)
            workbook.save(path_to_excel)
    except:
        bot.reply_to(message, f"Đã xảy ra lỗi trong thực thi log")
        



@bot.message_handler(commands=['dulieu'])
def send_data_command(message):
    try:
        path_to_txt = 'du_lieu.txt'
        path_to_excel = 'du_lieu.xlsx'

      
        with open(path_to_txt, 'r', encoding='utf-8') as file:
            data = file.readlines()

    
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for row_num, row_data in enumerate(data, start=1):
            sheet.cell(row=row_num, column=1, value=row_data.strip())

       
        workbook.save(path_to_excel)

       
        with open(path_to_excel, 'rb') as file:
            bot.send_document(message.chat.id, file)

    except Exception as e:
        bot.reply_to(message, f"Có lỗi xảy ra: {str(e)}")


@bot.message_handler(func=lambda msg: True)
def echo_all(message):
    bot.reply_to(message, message.text)

bot.polling()
